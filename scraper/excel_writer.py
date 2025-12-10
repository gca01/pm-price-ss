"""
Excel writer for Polymarket NBA price data.

Structure:
- One sheet per date (e.g., "2025-12-08")
- Games arranged horizontally (each game in its own column group)
- Screenshots stacked vertically for each game as time progresses

Layout example for a sheet:
    |  Col A-B (Game 1)  |  Col C-D (Game 2)  |  Col E-F (Game 3)  |
Row 1: SAC @ IND           PHX @ MIN            MIA @ ORL
Row 2: 1:00 PM / 2025-12-08  1:30 PM / 2025-12-08  2:00 PM / 2025-12-08
Row 3: https://polymarket... https://polymarket... https://polymarket...
Row 4: [Screenshot 1]      [Screenshot 1]       [Screenshot 1]
Row 5: Lows: SAC 25¢ / IND 40¢  Lows: PHX 20¢ / MIN 30¢  ...
Row 6: Captured: 06:00 AM  Captured: 06:00 AM   Captured: 06:00 AM
Row 7: (blank)             (blank)              (blank)
Row 8: [Screenshot 2]      [Screenshot 2]       [Screenshot 2]
Row 9: Lows: SAC 22¢ / IND 38¢  Lows: PHX 18¢ / MIN 28¢  ...
Row 10: Captured: 07:00 AM - FINAL  Captured: 07:00 AM  Captured: 07:00 AM
...
"""

from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Dict, Tuple
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from .config import EXCEL_FILE_PATH
from .utils import get_today_date_str, get_eastern_now, log_info, log_success, log_error, log_warning
from .game_screenshotter import GameScreenshotResult

# Layout constants
COLUMNS_PER_GAME = 3  # Each game takes 2 columns + 1 spacer column
HEADER_ROWS = 3  # Row 1: Game title, Row 2: Start time, Row 3: URL
ROWS_PER_ENTRY = 4  # Each hourly entry: Screenshot, Low prices, Capture time, blank row
IMAGE_WIDTH = 350  # Screenshot width in pixels for Excel
IMAGE_HEIGHT = 150  # Screenshot height in pixels for Excel


@dataclass
class GameState:
    """State of a game in the Excel sheet."""
    game_id: str
    url: Optional[str]
    is_final: bool
    column: int


def get_games_from_sheet(ws) -> Dict[str, GameState]:
    """
    Get all games from a worksheet with their current state.

    Args:
        ws: Worksheet to read from

    Returns:
        Dict mapping game_id to GameState (url, is_final, column)
    """
    games = {}

    col = 1
    while col <= ws.max_column:
        title_cell = ws.cell(row=1, column=col).value
        if not title_cell:
            col += COLUMNS_PER_GAME
            continue

        # Extract game_id from title (format: "Away @ Home")
        # We need to look at the date in row 2 to construct full game_id
        time_cell = ws.cell(row=2, column=col).value
        url_cell = ws.cell(row=3, column=col).value

        # Parse date from time cell (format: "HH:MM AM/PM / YYYY-MM-DD")
        game_date = None
        if time_cell and "/" in str(time_cell):
            parts = str(time_cell).split("/")
            if len(parts) >= 2:
                game_date = parts[-1].strip()

        # Parse teams from title (format: "Away @ Home")
        away_team = None
        home_team = None
        if title_cell and "@" in str(title_cell):
            parts = str(title_cell).split("@")
            if len(parts) == 2:
                away_team = parts[0].strip()
                home_team = parts[1].strip()

        if game_date and away_team and home_team:
            game_id = f"{game_date}_{away_team}_{home_team}"

            # Check if the game is final by looking at the last timestamp entry
            is_final = is_game_final(ws, col)

            games[game_id] = GameState(
                game_id=game_id,
                url=str(url_cell) if url_cell else None,
                is_final=is_final,
                column=col
            )

        col += COLUMNS_PER_GAME

    return games


def is_game_final(ws, game_col: int) -> bool:
    """
    Check if a game's last entry is marked as Final.

    Args:
        ws: Worksheet
        game_col: Column number for the game

    Returns:
        True if the last entry contains "FINAL"
    """
    # Find the last entry row for this game
    # Entry structure: Screenshot (row), Low prices (row+1), Timestamp (row+2), Blank (row+3)
    row = HEADER_ROWS + 1
    last_timestamp_row = None

    while row <= ws.max_row:
        # Timestamp is at row + 2 (after screenshot and low prices)
        timestamp_cell = ws.cell(row=row + 2, column=game_col).value
        if timestamp_cell and "Captured" in str(timestamp_cell):
            last_timestamp_row = row + 2
        row += ROWS_PER_ENTRY

    if last_timestamp_row:
        timestamp_value = ws.cell(row=last_timestamp_row, column=game_col).value
        if timestamp_value and "FINAL" in str(timestamp_value).upper():
            return True

    return False


def get_or_create_workbook(filepath: Path) -> Workbook:
    """Load existing workbook or create a new one."""
    if filepath.exists():
        return load_workbook(filepath)
    return Workbook()


def get_or_create_date_sheet(wb: Workbook, date_str: str):
    """Get or create a sheet for the given date."""
    # Clean sheet name (Excel doesn't allow certain characters)
    sheet_name = date_str  # Format: YYYY-MM-DD

    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    # Create new sheet
    ws = wb.create_sheet(title=sheet_name)

    # Remove default sheet if it exists and is empty
    if "Sheet" in wb.sheetnames:
        default_sheet = wb["Sheet"]
        if default_sheet.max_row == 1 and default_sheet.max_column == 1:
            wb.remove(default_sheet)

    return ws


def find_game_column(ws, game_id: str) -> Optional[int]:
    """
    Find the column where a game is located.

    Args:
        ws: Worksheet
        game_id: Game ID to find (format: "YYYY-MM-DD_Away_Home")

    Returns:
        Column number (1-based) or None if not found
    """
    # Parse game_id to get away and home teams
    # Format: "YYYY-MM-DD_Away_Home"
    parts = game_id.split("_")
    if len(parts) < 3:
        return None

    away_team = parts[1]
    home_team = parts[2]
    expected_title = f"{away_team} @ {home_team}"

    # Check row 1 for matching game title
    for col in range(1, ws.max_column + 1, COLUMNS_PER_GAME):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value and str(cell_value).strip() == expected_title:
            return col
    return None


def get_next_game_column(ws) -> int:
    """Get the next available column for a new game."""
    if ws.max_column <= 1 and ws.cell(row=1, column=1).value is None:
        return 1

    # Find the next empty column group
    col = 1
    while ws.cell(row=1, column=col).value is not None:
        col += COLUMNS_PER_GAME
    return col


def get_next_entry_row(ws, game_col: int) -> int:
    """
    Get the next available row for a new entry in a game column.

    Args:
        ws: Worksheet
        game_col: Column number for the game

    Returns:
        Row number for the next entry
    """
    # Start after header rows
    row = HEADER_ROWS + 1

    # Find the next empty entry slot
    # Entry structure: Screenshot (row), Low prices (row+1), Timestamp (row+2), Blank (row+3)
    # Check the timestamp row (row + 2) since screenshot row may have None value
    # even when an image is embedded there
    while ws.cell(row=row + 2, column=game_col).value is not None:
        row += ROWS_PER_ENTRY

    return row


def setup_game_header(ws, game_col: int, result: GameScreenshotResult):
    """
    Set up the header rows for a new game.

    Args:
        ws: Worksheet
        game_col: Starting column for this game
        result: GameScreenshotResult with game info
    """
    game = result.game

    # Style definitions
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    white_font = Font(bold=True, size=12, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row 1: Game title (Away @ Home)
    title_cell = ws.cell(row=1, column=game_col)
    title_cell.value = f"{game.away} @ {game.home}"
    title_cell.font = white_font
    title_cell.fill = header_fill
    title_cell.alignment = center_align
    title_cell.border = thin_border

    # Merge cells for title (only 2 columns, leave 3rd as spacer)
    ws.merge_cells(
        start_row=1, start_column=game_col,
        end_row=1, end_column=game_col + 1
    )

    # Row 2: Start time and Date
    time_cell = ws.cell(row=2, column=game_col)
    time_cell.value = f"{game.start_time or 'TBD'} / {game.game_date}"
    time_cell.font = Font(size=10, italic=True)
    time_cell.alignment = center_align
    time_cell.border = thin_border

    # Merge cells for time (only 2 columns)
    ws.merge_cells(
        start_row=2, start_column=game_col,
        end_row=2, end_column=game_col + 1
    )

    # Row 3: URL
    url_cell = ws.cell(row=3, column=game_col)
    url_cell.value = game.url or ""
    url_cell.font = Font(size=8, color="0066CC", underline="single")
    url_cell.alignment = center_align

    # Merge cells for URL (only 2 columns)
    ws.merge_cells(
        start_row=3, start_column=game_col,
        end_row=3, end_column=game_col + 1
    )

    # Set column widths (2 content columns + 1 narrow spacer)
    ws.column_dimensions[get_column_letter(game_col)].width = 25
    ws.column_dimensions[get_column_letter(game_col + 1)].width = 25
    ws.column_dimensions[get_column_letter(game_col + 2)].width = 3  # Spacer column


def add_entry_to_game(ws, game_col: int, entry_row: int, result: GameScreenshotResult):
    """
    Add a single entry (screenshot + prices + timestamp) to a game column.

    Args:
        ws: Worksheet
        game_col: Starting column for this game
        entry_row: Row to start this entry
        result: GameScreenshotResult with data
    """
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row 1 of entry: Screenshot
    if result.screenshot_path and Path(result.screenshot_path).exists():
        try:
            img = XLImage(str(result.screenshot_path))
            # Scale image to fit
            img.width = IMAGE_WIDTH
            img.height = IMAGE_HEIGHT

            # Position the image
            cell_ref = f"{get_column_letter(game_col)}{entry_row}"
            ws.add_image(img, cell_ref)

            # Set row height to accommodate image
            ws.row_dimensions[entry_row].height = IMAGE_HEIGHT * 0.75  # Convert to points

            log_success(f"Embedded screenshot at {cell_ref}")
        except Exception as e:
            log_warning(f"Could not embed image: {e}")
            # Fall back to path
            ws.cell(row=entry_row, column=game_col).value = str(result.screenshot_path)
    else:
        ws.cell(row=entry_row, column=game_col).value = "No screenshot"

    # Merge cells for screenshot row (only 2 columns, not the spacer)
    ws.merge_cells(
        start_row=entry_row, start_column=game_col,
        end_row=entry_row, end_column=game_col + 1
    )

    # Row 2 of entry: Low prices (two separate cells, not merged)
    low_row = entry_row + 1
    home_low = getattr(result, 'home_low_price', None)
    away_low = getattr(result, 'away_low_price', None)

    # Away team low in first column
    away_low_cell = ws.cell(row=low_row, column=game_col)
    if away_low is not None:
        away_low_cell.value = f"{result.game.away} Low: {away_low:.3f}"
    else:
        away_low_cell.value = f"{result.game.away} Low: -"
    away_low_cell.font = Font(size=9, color="CC6600")  # Orange for low prices
    away_low_cell.alignment = center_align

    # Home team low in second column
    home_low_cell = ws.cell(row=low_row, column=game_col + 1)
    if home_low is not None:
        home_low_cell.value = f"{result.game.home} Low: {home_low:.3f}"
    else:
        home_low_cell.value = f"{result.game.home} Low: -"
    home_low_cell.font = Font(size=9, color="CC6600")  # Orange for low prices
    home_low_cell.alignment = center_align

    # Row 3 of entry: Capture timestamp
    time_row = entry_row + 2
    timestamp = get_eastern_now().strftime("%I:%M %p")

    time_cell = ws.cell(row=time_row, column=game_col)
    # Check if this result is marked as final
    is_final = getattr(result, 'is_final', False)
    if is_final:
        time_cell.value = f"Captured: {timestamp} - FINAL"
        time_cell.font = Font(size=9, italic=True, bold=True, color="008000")  # Green for final
    else:
        time_cell.value = f"Captured: {timestamp}"
        time_cell.font = Font(size=9, italic=True, color="666666")
    time_cell.alignment = center_align

    # Merge timestamp cells (only 2 columns, not the spacer)
    ws.merge_cells(
        start_row=time_row, start_column=game_col,
        end_row=time_row, end_column=game_col + 1
    )

    # Row 4 of entry: blank row (spacer between entries)


def append_result(result: GameScreenshotResult, filepath: Optional[Path] = None) -> bool:
    """
    Append a single game result to the Excel file.

    Args:
        result: GameScreenshotResult to append
        filepath: Path to the Excel file. Uses default if None.

    Returns:
        True if append succeeded, False otherwise
    """
    if filepath is None:
        filepath = EXCEL_FILE_PATH

    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = get_or_create_workbook(filepath)
        ws = get_or_create_date_sheet(wb, result.game.game_date)

        # Find or create column for this game
        game_col = find_game_column(ws, result.game.game_id)

        if game_col is None:
            # New game - set up header
            game_col = get_next_game_column(ws)
            setup_game_header(ws, game_col, result)
            log_info(f"Created new game column at {game_col} for {result.game}")

        # Find next entry row for this game
        entry_row = get_next_entry_row(ws, game_col)

        # Add the entry
        add_entry_to_game(ws, game_col, entry_row, result)

        # Save
        wb.save(filepath)
        wb.close()

        log_success(f"Added entry for {result.game.game_id} at row {entry_row}")
        return True

    except Exception as e:
        log_error(f"Error appending to Excel: {e}")
        import traceback
        traceback.print_exc()
        return False


def append_results(results: List[GameScreenshotResult], filepath: Optional[Path] = None) -> int:
    """
    Append multiple game results to the Excel file.

    Args:
        results: List of GameScreenshotResult objects to append
        filepath: Path to the Excel file. Uses default if None.

    Returns:
        Number of successfully appended entries
    """
    if not results:
        log_info("No results to append")
        return 0

    if filepath is None:
        filepath = EXCEL_FILE_PATH

    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    appended = 0

    try:
        wb = get_or_create_workbook(filepath)

        for result in results:
            if not result.success:
                continue

            try:
                ws = get_or_create_date_sheet(wb, result.game.game_date)

                # Find or create column for this game
                game_col = find_game_column(ws, result.game.game_id)

                if game_col is None:
                    game_col = get_next_game_column(ws)
                    setup_game_header(ws, game_col, result)

                # Find next entry row
                entry_row = get_next_entry_row(ws, game_col)

                # Add the entry
                add_entry_to_game(ws, game_col, entry_row, result)

                appended += 1

            except Exception as e:
                log_error(f"Error adding {result.game}: {e}")

        # Save once at the end
        wb.save(filepath)
        wb.close()

        log_success(f"Appended {appended} entries to Excel")
        return appended

    except Exception as e:
        log_error(f"Error appending results to Excel: {e}")
        return 0


def get_sheet_names(filepath: Optional[Path] = None) -> List[str]:
    """Get list of sheet names (dates) in the workbook."""
    if filepath is None:
        filepath = EXCEL_FILE_PATH

    filepath = Path(filepath)

    if not filepath.exists():
        return []

    try:
        wb = load_workbook(filepath)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception as e:
        log_error(f"Error getting sheet names: {e}")
        return []


def get_entry_count(filepath: Optional[Path] = None, date_str: Optional[str] = None) -> Dict[str, int]:
    """
    Get count of entries per game for a date.

    Args:
        filepath: Path to Excel file
        date_str: Date to check (uses today if None)

    Returns:
        Dict mapping game_id to entry count
    """
    if filepath is None:
        filepath = EXCEL_FILE_PATH
    if date_str is None:
        date_str = get_today_date_str()

    filepath = Path(filepath)
    counts = {}

    if not filepath.exists():
        return counts

    try:
        wb = load_workbook(filepath)

        if date_str not in wb.sheetnames:
            wb.close()
            return counts

        ws = wb[date_str]

        # Check each game column
        col = 1
        while col <= ws.max_column:
            game_title = ws.cell(row=1, column=col).value
            if game_title:
                # Count entries for this game
                entry_count = 0
                row = HEADER_ROWS + 1
                while row <= ws.max_row:
                    if ws.cell(row=row + 1, column=col).value:  # Check timestamp row
                        entry_count += 1
                    row += ROWS_PER_ENTRY

                counts[str(game_title)] = entry_count

            col += COLUMNS_PER_GAME

        wb.close()
        return counts

    except Exception as e:
        log_error(f"Error getting entry count: {e}")
        return counts


def get_existing_games(filepath: Optional[Path] = None, date_str: Optional[str] = None) -> Dict[str, GameState]:
    """
    Get all existing games from the Excel file for a specific date.

    Args:
        filepath: Path to Excel file
        date_str: Date to check (uses today if None)

    Returns:
        Dict mapping game_id to GameState
    """
    if filepath is None:
        filepath = EXCEL_FILE_PATH
    if date_str is None:
        date_str = get_today_date_str()

    filepath = Path(filepath)

    if not filepath.exists():
        return {}

    try:
        wb = load_workbook(filepath)

        if date_str not in wb.sheetnames:
            wb.close()
            return {}

        ws = wb[date_str]
        games = get_games_from_sheet(ws)
        wb.close()
        return games

    except Exception as e:
        log_error(f"Error getting existing games: {e}")
        return {}
